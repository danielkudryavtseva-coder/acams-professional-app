"""Regenerate `src/app/data/portfolioHoldings.ts` from CAMS Portfolio.xlsx.

The frontend already has the live-data plumbing (FMP via
`usePortfolioLiveData`); this script only refreshes the static seed values
used as fallback display when the API is unavailable, plus the canonical
shares / cost basis / sector mapping.

Run: `py -3 scripts/extract_portfolio.py`
"""

from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

import openpyxl

SRC = Path(r"C:\Users\whitf\Downloads\CAMS Portfolio.xlsx")
OUT = Path(
    r"C:\Users\whitf\OneDrive\Desktop\DK_Local\professional-app\src\app\data\portfolioHoldings.ts"
)

# Order of rows in the "Long Equities" table (rows 7-33), reverse-engineered
# by matching purchase price + cached price to each per-ticker sheet.
EQUITY_ORDER = [
    "VRTX", "GS", "AMZN", "APO", "LLY", "AAPL", "MA", "WM", "NVO", "CVX",
    "GTBIF", "WMS", "UBER", "ACN", "SG", "ISRG", "CVE", "UNM", "ULTA", "RDNT",
    "PYPL", "NVDA", "UNH", "LVMH", "CMG", "CCJ", "VST",
]

# Use the canonical legal name where helpful; FMP returns its own version too.
NAME = {
    "AAPL": "Apple Inc.",
    "ACN": "Accenture plc",
    "AMZN": "Amazon.com, Inc.",
    "APO": "Apollo Global Management",
    "CCJ": "Cameco Corporation",
    "CMG": "Chipotle Mexican Grill",
    "CVE": "Cenovus Energy",
    "CVX": "Chevron Corporation",
    "GS": "The Goldman Sachs Group, Inc.",
    "GTBIF": "Green Thumb Industries Inc.",
    "ISRG": "Intuitive Surgical, Inc.",
    "LLY": "Eli Lilly and Company",
    "LVMH": "LVMH Moet Hennessy Louis Vuitton SE",
    "MA": "Mastercard Incorporated",
    "NVDA": "NVIDIA Corporation",
    "NVO": "Novo Nordisk A/S",
    "PYPL": "PayPal Holdings, Inc.",
    "RDNT": "RadNet, Inc.",
    "SG": "Sweetgreen, Inc.",
    "UBER": "Uber Technologies, Inc.",
    "ULTA": "Ulta Beauty, Inc.",
    "UNH": "UnitedHealth Group Incorporated",
    "UNM": "Unum Group",
    "VRTX": "Vertex Pharmaceuticals Incorporated",
    "VST": "Vistra Corp.",
    "WM": "Waste Management, Inc.",
    "WMS": "Advanced Drainage Systems, Inc.",
}

# Map workbook tickers -> the symbol used on the frontend (FMP override map
# in portfolioHoldings.ts handles converting to FMP's expected format).
FRONTEND_TICKER = {
    "LVMH": "MC",
}


def round2(x):
    if x is None:
        return None
    return round(float(x), 2)


def to_iso(d):
    if isinstance(d, (datetime, date)):
        return d.strftime("%Y-%m-%d")
    return None


FUND_NAME = {
    "GHYIX": "Goldman Sachs High Yield Municipal Fund",
    "VEMBX": "Vanguard Emerging Markets Bond Fund",
    "CQQQ": "Invesco China Technology ETF",
    "JPST": "JPMorgan Ultra-Short Income ETF",
}


def main() -> None:
    wb = openpyxl.load_workbook(SRC, data_only=True)
    portfolio = wb["Portfolio"]

    rows = []
    for i, sheet_ticker in enumerate(EQUITY_ORDER):
        r = 7 + i
        category = portfolio.cell(row=r, column=3).value
        shares = portfolio.cell(row=r, column=4).value
        avg_cost = portfolio.cell(row=r, column=5).value
        current_price = portfolio.cell(row=r, column=6).value

        ticker = FRONTEND_TICKER.get(sheet_ticker, sheet_ticker)
        current_value = round(shares * current_price)
        unrealized_gain = round(shares * (current_price - avg_cost))
        change_pct = round((current_price - avg_cost) / avg_cost * 100, 2) if avg_cost else 0

        rows.append({
            "ticker": ticker,
            "name": NAME.get(sheet_ticker, sheet_ticker),
            "shares": shares,
            "avgCost": round2(avg_cost),
            "currentPrice": round2(current_price),
            "sector": category,
            "currentValue": current_value,
            "unrealizedGain": unrealized_gain,
            "changePct": change_pct,
        })

    # Bonds — sourced from the Portfolio summary (rows 50-52) so totals
    # match the spreadsheet's roll-up numbers.
    bonds = []
    for r in range(50, 60):
        name = portfolio.cell(row=r, column=2).value
        shares = portfolio.cell(row=r, column=4).value
        if not isinstance(name, str) or not isinstance(shares, (int, float)):
            continue
        avg_cost = portfolio.cell(row=r, column=5).value
        current_price = portfolio.cell(row=r, column=6).value
        value = portfolio.cell(row=r, column=10).value
        # Normalize the bond name (replace lossy unicode in source).
        clean_name = (
            name.replace("4\xb0 ", "4.5% ")
            .replace("5\xb0 ", "5.5% ")
            .replace("6\xb0 ", "6.25% ")
            .replace("\xb0", "%")
        )
        bonds.append({
            "name": clean_name,
            "shares": shares,
            "purchasePrice": round2(avg_cost),
            "currentPrice": round2(current_price),
            "value": round(value) if isinstance(value, (int, float)) else None,
            "unrealizedGain": (
                round(shares * (current_price - avg_cost))
                if all(isinstance(x, (int, float)) for x in (shares, avg_cost, current_price))
                else None
            ),
        })

    # Mutual Funds sheet
    mfunds = []
    mf_ws = wb["Mutual Funds"]
    for r in range(5, mf_ws.max_row + 1):
        ticker = mf_ws.cell(row=r, column=3).value
        if not isinstance(ticker, str):
            continue
        purchase_price = mf_ws.cell(row=r, column=7).value
        current_price = mf_ws.cell(row=r, column=8).value
        units = mf_ws.cell(row=r, column=9).value
        purchase_value = mf_ws.cell(row=r, column=10).value
        current_value = mf_ws.cell(row=r, column=11).value
        if not all(isinstance(x, (int, float)) for x in (purchase_price, current_price, units)):
            continue
        unrealized = current_value - purchase_value if isinstance(current_value, (int, float)) and isinstance(purchase_value, (int, float)) else None
        mfunds.append({
            "ticker": ticker,
            "name": FUND_NAME.get(ticker, ticker),
            "shares": units,
            "purchasePrice": round2(purchase_price),
            "currentPrice": round2(current_price),
            "value": round(current_value) if isinstance(current_value, (int, float)) else None,
            "unrealizedGain": round(unrealized) if isinstance(unrealized, (int, float)) else None,
        })

    # Cash from CASH sheet (largest numeric value).
    cash = 0
    if "CASH" in wb.sheetnames:
        cws = wb["CASH"]
        for row in cws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, (int, float)) and cell > 1000:
                    cash = max(cash, cell)
    cash = round(cash)

    as_of = datetime.now().strftime("%Y-%m-%d")

    body_lines = [
        "  // Seed values mirror `CAMS Portfolio.xlsx` cached closes.",
        f"  // As-of: {as_of}. Run `py -3 scripts/extract_portfolio.py` to refresh.",
    ]
    for row in rows:
        ordered = {
            "ticker": row["ticker"],
            "name": row["name"],
            "shares": row["shares"],
            "avgCost": row["avgCost"],
            "currentPrice": row["currentPrice"],
            "sector": row["sector"],
            "currentValue": row["currentValue"],
            "unrealizedGain": row["unrealizedGain"],
            "changePct": row["changePct"],
        }
        parts = []
        for k, v in ordered.items():
            if isinstance(v, str):
                parts.append(f'{k}: {json.dumps(v)}')
            else:
                parts.append(f'{k}: {v}')
        body_lines.append("  { " + ", ".join(parts) + " },")

    bond_lines = []
    for b in bonds:
        parts = []
        for k, v in b.items():
            parts.append(f'{k}: {json.dumps(v) if isinstance(v, str) else v}')
        bond_lines.append("  { " + ", ".join(parts) + " },")

    fund_lines = []
    for f in mfunds:
        parts = []
        for k, v in f.items():
            parts.append(f'{k}: {json.dumps(v) if isinstance(v, str) else v}')
        fund_lines.append("  { " + ", ".join(parts) + " },")

    ts = (
        "// AUTO-GENERATED FROM `CAMS Portfolio.xlsx`.\n"
        "// Run `py -3 scripts/extract_portfolio.py` to regenerate.\n"
        "//\n"
        "// `currentPrice`/`currentValue`/`changePct` are seed values used as\n"
        "// fallback display when the FMP plug-in is unavailable. Add\n"
        "// `VITE_FMP_API_KEY` to `.env.local` to fetch live quotes.\n"
        "\n"
        "export interface PortfolioHolding {\n"
        "  ticker: string;\n"
        "  name: string;\n"
        "  shares: number;\n"
        "  avgCost: number;\n"
        "  currentPrice: number;\n"
        "  sector: string;\n"
        "  currentValue: number;\n"
        "  unrealizedGain: number;\n"
        "  changePct: number;\n"
        "}\n"
        "\n"
        "export interface FixedIncomePosition {\n"
        "  name: string;\n"
        "  shares: number;\n"
        "  purchasePrice: number;\n"
        "  currentPrice: number;\n"
        "  value: number;\n"
        "  unrealizedGain: number;\n"
        "}\n"
        "\n"
        "export interface MutualFundPosition {\n"
        "  ticker: string;\n"
        "  name: string;\n"
        "  shares: number;\n"
        "  purchasePrice: number;\n"
        "  currentPrice: number;\n"
        "  value: number;\n"
        "  unrealizedGain: number;\n"
        "}\n"
        "\n"
        "/**\n"
        " * Some portfolio tickers map to a different symbol on Financial Modeling Prep.\n"
        " * Anything not listed here is sent to FMP as-is.\n"
        " *\n"
        " * Known caveats on the free tier:\n"
        " * - `GTBIF` (Green Thumb, OTC) usually returns no data; falls back to seed values.\n"
        " * - `MC` is LVMH on Euronext Paris -> FMP wants `MC.PA`.\n"
        " */\n"
        "export const FMP_SYMBOL_OVERRIDES: Record<string, string> = {\n"
        '  MC: "MC.PA",\n'
        "};\n"
        "\n"
        f'export const PORTFOLIO_AS_OF = "{as_of}";\n'
        f"export const PORTFOLIO_CASH = {cash};\n"
        "\n"
        "export const PORTFOLIO_HOLDINGS: PortfolioHolding[] = [\n"
        + "\n".join(body_lines)
        + "\n];\n"
        "\n"
        "export const PORTFOLIO_BONDS: FixedIncomePosition[] = [\n"
        + "\n".join(bond_lines)
        + "\n];\n"
        "\n"
        "export const PORTFOLIO_MUTUAL_FUNDS: MutualFundPosition[] = [\n"
        + "\n".join(fund_lines)
        + "\n];\n"
    )

    OUT.write_text(ts, encoding="utf-8")
    eq_value = sum(r["currentValue"] for r in rows)
    eq_cost = sum(r["shares"] * r["avgCost"] for r in rows)
    bond_value = sum(b["value"] for b in bonds if isinstance(b["value"], (int, float)))
    bond_gain = sum(b["unrealizedGain"] for b in bonds if isinstance(b["unrealizedGain"], (int, float)))
    fund_value = sum(f["value"] for f in mfunds if isinstance(f["value"], (int, float)))
    fund_gain = sum(f["unrealizedGain"] for f in mfunds if isinstance(f["unrealizedGain"], (int, float)))
    total_value = eq_value + bond_value + fund_value
    total_gain = (eq_value - eq_cost) + bond_gain + fund_gain
    print(f"Wrote {OUT}")
    print(f"  equities:    {len(rows):3d}  value ${eq_value:>10,.0f}  gain ${eq_value - eq_cost:>+10,.0f}")
    print(f"  bonds:       {len(bonds):3d}  value ${bond_value:>10,.0f}  gain ${bond_gain:>+10,.0f}")
    print(f"  mutual fund: {len(mfunds):3d}  value ${fund_value:>10,.0f}  gain ${fund_gain:>+10,.0f}")
    print(f"  cash:                value ${cash:>10,.0f}")
    print(f"  ----")
    print(f"  TOTAL VALUE:        ${total_value:>10,.0f}")
    print(f"  TOTAL GAIN:         ${total_gain:>+10,.0f}")


if __name__ == "__main__":
    main()
